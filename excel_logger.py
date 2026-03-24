"""
excel_logger.py — Persistent Excel Trade Logger
ALL days in ONE file — appends across sessions, never resets.
Sheets: All Trades | QQQ | TQQQ | Performance | Equity Curve
"""

import os, threading, logging
from datetime import datetime
import pytz

log        = logging.getLogger(__name__)
IST        = pytz.timezone("Asia/Kolkata")
ET         = pytz.timezone("America/New_York")
EXCEL_PATH = "trades_log.xlsx"
COMMISSION = 0.0005
_lock      = threading.Lock()

C = {
    "bg":       "0D1117", "card":     "161B22", "alt":      "1C2333",
    "border":   "21262D", "green":    "00FF88", "red":      "FF4466",
    "yellow":   "F7C948", "purple":   "8B8BFF", "blue":     "4D9EFF",
    "text":     "E6EDF3", "dim":      "8B949E",
    "long_bg":  "0A2218", "short_bg": "221018",
    "flip_bg":  "22200A", "eod_bg":   "0A0A22",
    "win_bg":   "081A10", "loss_bg":  "1A0808",
}

HEADERS = [
    "Date","Time (ET)","Time (IST)","Symbol","Action","Side",
    "Entry ($)","VWAP ($)","Qty","Notional ($)",
    "Gross P&L ($)","Commission ($)","Net P&L ($)",
    "Cumul. Net P&L ($)","Equity ($)","Win/Loss","Notes"
]
WIDTHS = [12,10,10,8,18,8,11,11,9,13,13,12,12,16,14,10,22]


def _make_wb():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XS
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    wb   = Workbook()
    thin = XS(style="thin", color=C["border"])
    bdr  = lambda: Border(left=thin, right=thin, top=thin, bottom=thin)

    def fn(col=C["text"], bold=False, sz=9):
        return Font(name="Consolas", bold=bold, size=sz, color=col)
    def fp(col): return PatternFill("solid", fgColor=col)
    def fa(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def make_trades_sheet(ws, accent, tab_color, title):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab_color
        span = get_column_letter(len(HEADERS))

        ws.merge_cells(f"A1:{span}1")
        ws["A1"].value     = f"⚡  VWAP BOT — {title}"
        ws["A1"].font      = fn(accent, True, 13)
        ws["A1"].fill      = fp(C["bg"])
        ws["A1"].alignment = fa()
        ws.row_dimensions[1].height = 34

        ws.merge_cells(f"A2:{span}2")
        ws["A2"].value = (
            "Strategy: VWAP cross on 1-min candles  |  Long > VWAP  |  Short < VWAP  |"
            "  Flip on candle close  |  EOD 15:55 ET / 1:25 AM IST  |  Alpaca Paper Trading"
        )
        ws["A2"].font      = Font(name="Consolas", size=8, italic=True, color=C["dim"])
        ws["A2"].fill      = fp("0A0F14")
        ws["A2"].alignment = fa()
        ws.row_dimensions[2].height = 15

        for i,(h,w) in enumerate(zip(HEADERS, WIDTHS), 1):
            c = ws.cell(row=3, column=i)
            c.value=h; c.font=fn(accent, True, 9)
            c.fill=fp(C["alt"]); c.alignment=fa(wrap=True); c.border=bdr()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A4"
        return ws

    ws_all  = wb.active
    ws_all.title = "All Trades"
    make_trades_sheet(ws_all,  C["green"],  C["green"],  "ALL TRADES LOG")

    ws_qqq  = wb.create_sheet("QQQ")
    make_trades_sheet(ws_qqq,  C["blue"],   "4D9EFF",    "QQQ TRADES")

    ws_tqqq = wb.create_sheet("TQQQ")
    make_trades_sheet(ws_tqqq, C["purple"], "8B8BFF",    "TQQQ TRADES")

    # ── PERFORMANCE SHEET ──────────────────────────────────────────────────
    ws_p = wb.create_sheet("Performance")
    ws_p.sheet_view.showGridLines = False
    ws_p.sheet_properties.tabColor = C["yellow"]

    ws_p.merge_cells("A1:G1")
    ws_p["A1"].value     = "⚡  VWAP BOT — PERFORMANCE SUMMARY (Live Excel Formulas)"
    ws_p["A1"].font      = fn(C["yellow"], True, 13)
    ws_p["A1"].fill      = fp(C["bg"])
    ws_p["A1"].alignment = fa()
    ws_p.row_dimensions[1].height = 34

    ph = [("Metric",30),("QQQ",18),("TQQQ",18),("Combined",18),("Paper Target",18),("Notes",30)]
    for i,(h,w) in enumerate(ph,1):
        c = ws_p.cell(row=2,column=i)
        c.value=h; c.font=fn(C["yellow"],True,9); c.fill=fp(C["alt"])
        c.alignment=fa(wrap=True); c.border=bdr()
        ws_p.column_dimensions[get_column_letter(i)].width=w
    ws_p.row_dimensions[2].height = 26

    # All formulas reference column M (Net P&L) of All Trades
    perf = [
        ("Total Net P&L ($)",
         '=SUMIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"QQQ")',
         '=SUMIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"TQQQ")',
         "=B3+C3","—","After commission"),
        ("Total Commission ($)",
         '=SUMIFS(\'All Trades\'!L:L,\'All Trades\'!D:D,"QQQ")',
         '=SUMIFS(\'All Trades\'!L:L,\'All Trades\'!D:D,"TQQQ")',
         "=B4+C4","—","$0.0005/share"),
        ("Total Trades",
         '=COUNTIF(\'All Trades\'!D:D,"QQQ")',
         '=COUNTIF(\'All Trades\'!D:D,"TQQQ")',
         "=B5+C5","~22,000/yr","Zarattini Table 3"),
        ("Winning Trades",
         '=COUNTIFS(\'All Trades\'!D:D,"QQQ",\'All Trades\'!M:M,">"&0)',
         '=COUNTIFS(\'All Trades\'!D:D,"TQQQ",\'All Trades\'!M:M,">"&0)',
         "=B6+C6","—","Net P&L > 0"),
        ("Losing Trades",
         '=COUNTIFS(\'All Trades\'!D:D,"QQQ",\'All Trades\'!M:M,"<"&0)',
         '=COUNTIFS(\'All Trades\'!D:D,"TQQQ",\'All Trades\'!M:M,"<"&0)',
         "=B7+C7","—","Net P&L < 0"),
        ("Hit Ratio",
         "=IFERROR(B6/B5,0)","=IFERROR(C6/C5,0)",
         "=IFERROR((B6+C6)/(B5+C5),0)","~17%","Low is normal"),
        ("Avg Win ($)",
         '=IFERROR(AVERAGEIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"QQQ",\'All Trades\'!M:M,">"&0),0)',
         '=IFERROR(AVERAGEIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"TQQQ",\'All Trades\'!M:M,">"&0),0)',
         "=IFERROR((B3+C3)/(B6+C6),0)","—",""),
        ("Avg Loss ($)",
         '=IFERROR(AVERAGEIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"QQQ",\'All Trades\'!M:M,"<"&0),0)',
         '=IFERROR(AVERAGEIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"TQQQ",\'All Trades\'!M:M,"<"&0),0)',
         "=IFERROR((B3+C3)/(B7+C7),0)","—",""),
        ("Gain:Loss Ratio",
         "=IFERROR(ABS(B8/B9),0)","=IFERROR(ABS(C8/C9),0)",
         "=IFERROR(ABS(D8/D9),0)","~5.7x","Zarattini Table 4"),
        ("Max Win ($)",
         '=IFERROR(MAXIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"QQQ"),0)',
         '=IFERROR(MAXIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"TQQQ"),0)',
         "=MAX(B11,C11)","—",""),
        ("Max Loss ($)",
         '=IFERROR(MINIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"QQQ"),0)',
         '=IFERROR(MINIFS(\'All Trades\'!M:M,\'All Trades\'!D:D,"TQQQ"),0)',
         "=MIN(B12,C12)","—",""),
        ("Total Days Traded",
         '=IFERROR(SUMPRODUCT(1/COUNTIFS(\'All Trades\'!A:A,\'All Trades\'!A:A,\'All Trades\'!D:D,"QQQ")),0)',
         '=IFERROR(SUMPRODUCT(1/COUNTIFS(\'All Trades\'!A:A,\'All Trades\'!A:A,\'All Trades\'!D:D,"TQQQ")),0)',
         "=MAX(B13,C13)","—","Unique trading days"),
        ("Latest Equity ($)",
         "=IFERROR(LOOKUP(2,1/(\'All Trades\'!O:O<>\"\"),\'All Trades\'!O:O),0)",
         "=B14","=B14","—","Most recent equity"),
    ]

    for r,row in enumerate(perf,3):
        for ci,val in enumerate(row,1):
            c = ws_p.cell(row=r,column=ci,value=val)
            c.fill=fp(C["alt"] if r%2==0 else C["card"])
            c.alignment=fa() if ci>1 else Alignment(horizontal="left",vertical="center")
            c.border=bdr(); c.font=fn(C["text"],ci==1,9)
        ws_p.row_dimensions[r].height=22

    for col in (2,3,4): ws_p.cell(row=8,column=col).number_format="0.0%"

    # ── EQUITY CURVE SHEET ─────────────────────────────────────────────────
    ws_eq = wb.create_sheet("Equity Curve")
    ws_eq.sheet_view.showGridLines = False
    ws_eq.sheet_properties.tabColor = C["blue"]

    ws_eq.merge_cells("A1:D1")
    ws_eq["A1"].value     = "⚡  EQUITY CURVE — Daily Tracking"
    ws_eq["A1"].font      = fn(C["blue"], True, 13)
    ws_eq["A1"].fill      = fp(C["bg"])
    ws_eq["A1"].alignment = fa()
    ws_eq.row_dimensions[1].height = 34

    eq_hdrs = [("Date",14),("Equity ($)",16),("Daily P&L ($)",16),("Cumul. P&L ($)",18)]
    for i,(h,w) in enumerate(eq_hdrs,1):
        c = ws_eq.cell(row=2,column=i)
        c.value=h; c.font=fn(C["blue"],True,9); c.fill=fp(C["alt"])
        c.alignment=fa(wrap=True); c.border=bdr()
        ws_eq.column_dimensions[get_column_letter(i)].width=w
    ws_eq.row_dimensions[2].height=26
    ws_eq.freeze_panes="A3"

    wb.save(EXCEL_PATH)
    log.info(f"✅ Excel workbook created: {EXCEL_PATH}")


def _ensure():
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_PATH):
        _make_wb()
        return load_workbook(EXCEL_PATH)
    try:
        wb = load_workbook(EXCEL_PATH)
        for s in ("All Trades","QQQ","TQQQ","Performance","Equity Curve"):
            if s not in wb.sheetnames: raise ValueError(f"Missing: {s}")
        return wb
    except Exception as e:
        log.warning(f"Excel corrupt ({e}) — recreating")
        try: os.remove(EXCEL_PATH)
        except: pass
        _make_wb()
        return load_workbook(EXCEL_PATH)


def log_trade_to_excel(trade: dict, account_equity: float, cum_net_pnl: float):
    with _lock:
        try: _write_row(trade, account_equity, cum_net_pnl)
        except Exception as e: log.error(f"Excel write failed: {e}")


def _write_row(trade: dict, equity: float, cum_pnl: float):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XS

    wb = _ensure()

    action    = str(trade.get("action", ""))
    symbol    = str(trade.get("symbol", ""))
    price     = float(trade.get("price",  0) or 0)
    vwap_val  = float(trade.get("vwap",   0) or 0)
    qty       = int(trade.get("qty",      0) or 0)
    gross_pnl = trade.get("realized_pnl")
    ts_raw    = trade.get("time", "")

    try:
        ts = datetime.fromisoformat(str(ts_raw).replace("Z", "+00:00"))
        if ts.tzinfo is None: ts = ET.localize(ts)
        ts_et  = ts.astimezone(ET)
        ts_ist = ts.astimezone(IST)
    except Exception:
        ts_et = ts_ist = datetime.now(ET)

    date_str    = ts_et.strftime("%Y-%m-%d")
    time_et_str = ts_et.strftime("%H:%M:%S")
    time_ist_str= ts_ist.strftime("%H:%M:%S")

    side     = "LONG" if "LONG" in action else "SHORT" if "SHORT" in action else "—"
    notional = round(price * qty, 2)
    comm     = round(qty * COMMISSION, 2)
    gross_v  = float(gross_pnl) if gross_pnl is not None else None
    net_pnl  = round(gross_v - comm, 2) if gross_v is not None else None
    win_loss = ("WIN" if net_pnl and net_pnl > 0 else
                "LOSS" if net_pnl and net_pnl < 0 else "—")

    if   "ENTER" in action and "LONG"  in action: bg = C["long_bg"]
    elif "ENTER" in action and "SHORT" in action: bg = C["short_bg"]
    elif "FLIP"  in action:                        bg = C["flip_bg"]
    elif "EOD"   in action:                        bg = C["eod_bg"]
    else:                                           bg = C["card"]

    thin   = XS(style="thin", color=C["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ACTION_COLORS = {
        "ENTER LONG":   C["green"],  "ENTER SHORT":  C["red"],
        "FLIP → LONG":  C["yellow"], "FLIP → SHORT": C["yellow"],
        "EOD CLOSE":    C["purple"],
    }

    row_vals = [
        date_str, time_et_str, time_ist_str, symbol, action, side,
        price, vwap_val, qty, notional,
        gross_v  if gross_v  is not None else "",
        comm,
        net_pnl  if net_pnl  is not None else "",
        round(cum_pnl, 2), round(equity, 2),
        win_loss, ""
    ]

    target_sheets = ["All Trades"]
    if symbol in ("QQQ", "TQQQ"): target_sheets.append(symbol)

    for sname in target_sheets:
        if sname not in wb.sheetnames: continue
        ws       = wb[sname]
        next_row = max(ws.max_row + 1, 4)
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=next_row, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

            if ci in (11, 13, 14) and val != "":
                try:
                    col = C["green"] if float(val) >= 0 else C["red"]
                    cell.font = Font(name="Consolas", size=9, bold=True, color=col)
                except: cell.font = Font(name="Consolas", size=9, color=C["text"])
            elif ci == 5:
                cell.font = Font(name="Consolas", size=9, bold=True,
                                 color=ACTION_COLORS.get(action, C["text"]))
            elif ci == 6:
                col = C["green"] if val=="LONG" else C["red"] if val=="SHORT" else C["dim"]
                cell.font = Font(name="Consolas", size=9, bold=True, color=col)
            elif ci == 16:
                col = C["green"] if val=="WIN" else C["red"] if val=="LOSS" else C["dim"]
                cell.font = Font(name="Consolas", size=9, bold=True, color=col)
            elif ci == 17:
                cell.font = Font(name="Consolas", size=8, italic=True, color=C["dim"])
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = Font(name="Consolas", size=9, color=C["text"])
        ws.row_dimensions[next_row].height = 20

    # Update Equity Curve sheet on EOD
    if "EOD" in action:
        _update_equity_curve(wb, date_str, equity, cum_pnl)

    wb.save(EXCEL_PATH)
    log.info(f"Excel ✅ {symbol} {action} | net=${net_pnl} | equity=${equity:.0f}")


def _update_equity_curve(wb, date_str: str, equity: float, cum_pnl: float):
    """Add a row to Equity Curve sheet after each EOD."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XS

    ws  = wb["Equity Curve"]
    row = max(ws.max_row + 1, 3)

    # Daily P&L = difference from previous equity row
    prev_equity = 0.0
    if row > 3:
        try: prev_equity = float(ws.cell(row=row-1, column=2).value or 0)
        except: pass

    daily_pnl = equity - prev_equity if prev_equity > 0 else 0

    thin   = XS(style="thin", color="21262D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    vals = [date_str, round(equity,2), round(daily_pnl,2), round(cum_pnl,2)]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill   = PatternFill("solid", fgColor="0D1117" if row%2==0 else "161B22")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ci in (3, 4) and val != "":
            try:
                col = "00FF88" if float(val) >= 0 else "FF4466"
                cell.font = Font(name="Consolas", size=9, bold=True, color=col)
            except: cell.font = Font(name="Consolas", size=9, color="E6EDF3")
        else:
            cell.font = Font(name="Consolas", size=9, color="E6EDF3")
    ws.row_dimensions[row].height = 20
